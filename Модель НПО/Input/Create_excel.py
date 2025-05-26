import os, openpyxl, numpy as np, pandas as pd
from itertools import chain
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Правило условного форматирования для индивидуальных коэффициентов развития:
color_scale_rule = ColorScaleRule(
start_type="min", start_color="FAEEDD",  # Цвет бедра испуганной нимфы
end_type="max", end_color="A8E4A0"  # Цвет "Бабушкины яблоки"
)

def format_output(writer, save_dir):
    wb = load_workbook(writer)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        for row in sheet.iter_rows(min_row=2, values_only=False):  # min_row=2, чтобы пропустить заголовки
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
    wb.save(save_dir)
    wb.close()

# Функция Покраска ячеек треугольника:
def yellow_diag(my_filename, n):
    wb = load_workbook(my_filename)
    for i in range(1, n+1):
        wb['Треуг_коэф'].cell(row=n+2-i, column = i + 1).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    wb.save(my_filename)

def DPUf_results(CF, triang_df, future, ug, n, res_dir):
    with pd.ExcelWriter(res_dir + str(ug) + '_ДПУф' + '.xlsx', engine = 'openpyxl') as writer:    
        CF.to_excel(writer, sheet_name='ДПУф_Итог', index=False)
        ws_CF = writer.sheets['ДПУф_Итог']
        # Устанавливаем ширину и формат для столбцов
        for col in ws_CF.columns:
            max_length = 0
            column = col[0].column_letter  # Получаем букву столбца
            # Устанавливаем формат для ячеек
            if column in ['A', 'B', 'C', 'D', 'G', 'N']:
                for cell in col:
                    cell.number_format = '### ### ### ### ##0'
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            elif column == 'F':
                for cell in col:
                    cell.number_format = '##0.000'
            elif column in ['L', 'M']:
                for cell in col:
                    cell.number_format = '##0.00000'
            # Устанавливаем ширину столбца
            ws_CF.column_dimensions[column].width = 15 if column != 'E' else 12
        triang_df.to_excel(writer, sheet_name='Треуг_коэф', index=False)
        ws_triang = writer.sheets['Треуг_коэф']
        # Устанавливаем ширину и формат для столбцов
        for col in ws_triang.columns:
            column = col[0].column_letter
            list_cols = [get_column_letter(i) for i in range(n + 2, n + 5)] + [get_column_letter(i) for i in range(n + 8, n + 10)]
            for cell in col:
                cell.number_format = '### ### ### ### ##0'
                if column in list_cols:
                    cell.number_format = '##0.00000'
            ws_triang.column_dimensions[column].width = 15
        # Записываем третий DataFrame
        future.to_excel(writer, sheet_name='ДПУф_будущие', index=False)
        ws_future = writer.sheets['ДПУф_будущие']
        # Устанавливаем ширину и формат для столбцов
        for col in ws_future.columns:
            column = col[0].column_letter
            for cell in col:
                cell.number_format = '### ### ### ### ##0'
            ws_future.column_dimensions[column].width = 14
    yellow_diag(res_dir + str(ug) + '_ДПУф' + '.xlsx', n)

def ORSf_results(CF_subr, triang_orsf, subr_future, ug, n, res_dir):
    with pd.ExcelWriter(res_dir + str(ug) + '_ОРСф' + '.xlsx', engine = 'openpyxl') as writer:    
        CF_subr.to_excel(writer, sheet_name='ОРСф_Итог', index=False)
        ws_CF = writer.sheets['ОРСф_Итог']
        # Устанавливаем ширину и формат для столбцов
        for col in ws_CF.columns:
            max_length = 0
            column = col[0].column_letter  # Получаем букву столбца
            # Устанавливаем формат для ячеек
            if column in ['A', 'B', 'C', 'D', 'G', 'N']:
                for cell in col:
                    cell.number_format = '### ### ### ### ##0'
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            elif column == 'F':
                for cell in col:
                    cell.number_format = '##0.000'
            elif column in ['L', 'M']:
                for cell in col:
                    cell.number_format = '##0.00000'
            # Устанавливаем ширину столбца
            ws_CF.column_dimensions[column].width = 15 if column != 'E' else 12
        triang_orsf.to_excel(writer, sheet_name='Треуг_коэф', index=False)
        ws_triang = writer.sheets['Треуг_коэф']
        # Устанавливаем ширину и формат для столбцов
        for col in ws_triang.columns:
            column = col[0].column_letter
            list_cols = [get_column_letter(i) for i in range(n + 2, n + 5)] + [get_column_letter(i) for i in range(n + 8, n + 10)]
            for cell in col:
                cell.number_format = '### ### ### ### ##0'
                if column in list_cols:  # Пример: для столбцов F, G, H
                    cell.number_format = '##0.00000'
            ws_triang.column_dimensions[column].width = 15
        # Записываем третий DataFrame
        subr_future.to_excel(writer, sheet_name='ОРСф_будущие', index=False)
        ws_future = writer.sheets['ОРСф_будущие']
        # Устанавливаем ширину и формат для столбцов
        for col in ws_future.columns:
            column = col[0].column_letter
            for cell in col:
                cell.number_format = '### ### ### ### ##0'
            ws_future.column_dimensions[column].width = 14
    yellow_diag(res_dir + str(ug) + '_ОРСф' + '.xlsx', n)

# Функция сохранения результатов расчёта в excel-файлик:
def create_excel_results(root_res, CF, ind_coef, coefs_df, triang_df, select_LR, future, Cj, ug, n, tr_type='paid', res = 'ibnr'):
    # Создание/открытие файла:
    my_filename = root_res + str(ug) + '_' + res + '.xlsx'
    if os.path.exists(my_filename):
        if tr_type=='paid':
            os.remove(my_filename)
            wb = Workbook()
            wb.remove(wb.worksheets[0])
        else:
            wb = load_workbook(my_filename)
    else:
        wb = Workbook()
        wb.remove(wb.worksheets[0])
    # Запись данных в ексель-файл:
    my_dfs = list(zip([ind_coef, select_LR, triang_df, future, CF], ['ДПУВно_коэф', 'Выбор_убыт', 'ДПУВно_треуг', 'ДПУВно_будущие', 'ДПУВно_итог']))
    for i in range(len(my_dfs)):
        ws = wb.create_sheet(my_dfs[i][1] + '_' + tr_type)
        for r in dataframe_to_rows(my_dfs[i][0], index=False, header=True):
            ws.append(r)
        globals()['ws{}'.format(i+1)] = ws # Создание глобальных переменных соответствующих каждому листу
        ws.sheet_properties.tabColor = 'FFDEAD' if tr_type == 'paid' else 'BDECB6' # Цвет листа
    # Таблицы вариантов коэффициентов на лист ДПУВно_треуг:
    for r_idx, r in enumerate(dataframe_to_rows(coefs_df, index=False, header=False), start=n+3):
        for c_idx, value in enumerate(r, start=1):
            ws3.cell(row=r_idx, column=c_idx, value=value)
            ws3.cell(row=r_idx, column=c_idx).number_format = '# ##0.000000'
    # Присвоение форматов числовых ячеек:
    for i in range(2, n+2):
        for j in chain(range(1, n+1), range(n+4, n+7), range(n+8, n+14)):
            ws3.cell(row=i, column=j+1).number_format = '### ### ### ##0'
    for i in range(2, n+2):
        for j in chain(range(n+1, n+4), range(n+7, n+8)):
            ws3.cell(row=i, column=j+1).number_format = '# ##0.00000'
    for row in ws4.iter_rows():
        for cell in row:
            cell.number_format = '### ### ### ##0'
    for ws in [ws1, ws2]:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.number_format = '# ##0.000000'
    for row in ws5['E2:E' + str(ws5.max_row)]:
        for cell in row:
            cell.number_format = '# ##0.00000'
    for diap in ['C2:C', 'H2:H', 'M2:M']:
        for row in ws5[diap + str(ws5.max_row)]:
            for cell in row:
                cell.number_format = '### ### ### ##0'
    # Покраска ячеек треугольника:
    for i in range(1, n+1):
        ws3.cell(row=n+2-i, column = i + 1).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    # Ширина столбцов:
    for ws in [ws2, ws3, ws4, ws5]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 12
    # Добавление правил условного форматирования для коэф развития и видов убыточности:
    for i in range(1, n-1):
        col_letter = get_column_letter(i)
        ws1.conditional_formatting.add(f'{col_letter}2:{col_letter}{n+1-i}', color_scale_rule)
    for i in range(2, n+2):
        ws2.conditional_formatting.add(f'A{i}:F{i}', color_scale_rule)
    for i in range(1, 19):
        col_letter = get_column_letter(i)
        ws3.conditional_formatting.add(f'{col_letter}{n+4}:{col_letter}{n+9}', color_scale_rule)
    # Добавление треугольника выбранных коэффициентов:
    for col in range (n+1, n+1+n):
        for row in range(2, n+2+n-col):
            ws1.cell(column = col, row = row, value=Cj[col-n-1])
    # Создание графиков коэффициентов:
    for col in range(1, n):
        chart_b = BarChart()
        series_b = Reference(ws3, min_col=col+1, max_col=col+1, min_row=3,  max_row = n+2-col)
        chart_b.series.append(Series(series_b, title=f"Вес"))
        chart_b.y_axis.majorGridlines = None
        chart = LineChart()
        series_1 = Reference(ws1, min_col=col, max_col=col, min_row=2,  max_row = n+1-col)
        series_2 = Reference(ws1, min_col=col+n, max_col=col+n, min_row=2,  max_row = n+1-col)
        chart.series.append(Series(series_1, title=f"Динамика инд коэф"))
        chart.series.append(Series(series_2, title=f"Выбранный коэф"))
        chart.y_axis.axId = 200
        chart_b.y_axis.crosses = "max"
        chart_b += chart
        chart_b.title = "Динамика коэф " + str(col)
        chart_b.width = 23
        ws1.add_chart(chart_b, f'{get_column_letter(15 - 14 * (col%2))}{n+3 + 15*((col-1)//2)}')
    # Создание графиков убыточностей:
    chart = LineChart()
    for col in range(1, 7):
        series = Reference(ws2, min_col=col, max_col=col, min_row=2,  max_row = n+1)
        chart.series.append(Series(series, title=ws2.cell(row=1, column=col).value))
    chart.title = "Динамика убыточности"
    chart_b.width = 35
    ws2.add_chart(chart, 'H2')
    wb.save(my_filename)
    
def compare_paid_and_increment(root_res, dev_coef_paid, dev_coef_inc, ug, res='ibnr'):
    my_filename = root_res + str(ug) + '_' + res + '.xlsx'
    wb = load_workbook(my_filename)
    ws = wb.create_sheet('Koef_deviation')
    coefs_df = pd.DataFrame([dev_coef_paid, dev_coef_inc])
    coefs_df.insert(0, 'Тип коэфф', ['dev_coef_paid', 'dev_coef_inc'])
    for r in dataframe_to_rows(coefs_df, index=False, header=True): # Запись данных в файл
            ws.append(r)
    for row in ws.iter_rows(min_row=2): # Форматирование ячеек
            for cell in row:
                cell.number_format = '# ##0.000000'
    wb.move_sheet(ws, offset=-len(wb.sheetnames)+1)
    ws.sheet_properties.tabColor = '7FFFD4' # Цвет листа
    wb.save(my_filename)